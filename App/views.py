from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.shortcuts import render, redirect
from django.contrib.auth.views import LoginView, PasswordResetView, PasswordChangeView, PasswordResetConfirmView
from App.forms import  *
from django.contrib.auth import logout
from .models import Profesor, SemestreAcademico, Profile, Semestre_Academico_Profesores, Sustentacion, Estudiante, Cursos_Grupos, Curso, Grupo
from django.contrib.admin.views.decorators import staff_member_required
from .forms import ExcelUploadForm
from django.contrib.auth.models import User
from django.db import transaction, DatabaseError
import pandas as pd
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
import time
from django.db import connection
from django.utils.dateparse import parse_datetime
from django.contrib.auth.decorators import login_required
# views.py
from .AlgorimoGenetico import generar_horarios, guardar_horario
from .utils import send_whatsapp_message

from django.views.decorators.http import require_http_methods

from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.db.models import F

from django.utils.timezone import localtime

from django.utils.timezone import localtime
from datetime import timedelta
from babel.dates import format_date
# Establecer el locale a español
@staff_member_required
@csrf_exempt
@require_http_methods(["GET", "POST"])
def ejecutar_algoritmo(request):
    semestres = SemestreAcademico.objects.all()
    semestre_vigente = SemestreAcademico.objects.filter(vigencia=True).first()
    if not semestre_vigente:
        print("No hay semestre vigente.")
        return []  # O alguna otra acción en caso de no haber semestre vigente

    cursos_grupos = Cursos_Grupos.objects.filter(semestre=semestre_vigente)
    semestre_id = request.GET.get('semestre')
    tipo_sustentacion = request.GET.get('tipo_sustentacion')
    curso_grupo_id = request.GET.get('curso_grupo')

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            tipo_sustentacion = body.get('tipo_sustentacion')
        except json.JSONDecodeError:
            tipo_sustentacion = request.POST.get('tipo_sustentacion')

        print("Tipo de sustentación:", tipo_sustentacion)

        if tipo_sustentacion not in ['FINAL', 'PARCIAL']:
            messages.error(request, "Seleccione el tipo de sustentación.")
            return redirect('ejecutar_algoritmo')

        semestre_actual = SemestreAcademico.objects.filter(vigencia=True).first()

        if not semestre_actual:
            messages.error(request, "No hay un semestre en vigencia.")
            return redirect('ejecutar_algoritmo')

        profesores_semestre_actual = Profesor.objects.filter(
            cursos_grupos__semestre=semestre_actual
        ).distinct()

        profesores_sin_disponibilidad = []

        for profesor in profesores_semestre_actual:
            disponibilidad = Profesores_Semestre_Academico.objects.filter(
                profesor=profesor,
                semestre=semestre_actual
            )
            if not disponibilidad.exists():
                profesores_sin_disponibilidad.append(profesor)

        if profesores_sin_disponibilidad:
            context = {
                'profesores_sin_disponibilidad': profesores_sin_disponibilidad,
                'semestres': semestres,
                'cursos_grupos': cursos_grupos,
                'semestre_id': semestre_id,
                'tipo_sustentacion': tipo_sustentacion,
                'curso_grupo_id': curso_grupo_id,
            }
            return render(request, 'admin/ejecutar_algoritmo.html', context)

        if tipo_sustentacion == 'FINAL':
            sustentaciones_vigentes = Sustentacion.objects.filter(
                cursos_grupos__semestre=semestre_actual
            )

            for sustentacion in sustentaciones_vigentes:
                if not (sustentacion.jurado1 and sustentacion.jurado2 and sustentacion.asesor):
                    messages.error(request, "Primero se deben de generar las sustentaciones parciales y que todas las sustentaciones tengan jurados asignados")
                    return redirect('ejecutar_algoritmo')
        try:
            mejor_horario = []
            for _ in range(4):  # Ejecutar 4 veces consecutivas
                mejor_horario = generar_horarios(tipo_sustentacion)
                guardar_horario(mejor_horario)  # Guardar el horario en la base de datos
        except Exception as e:
            messages.error(request, f"Error durante la ejecución del algoritmo: revisar que haya registrada semanas de sustentacion para este tipo de sustentacion en el semestre actual, tambien disponibilidad de los docentes, sustentaciones, etc")
            return redirect('ejecutar_algoritmo')

        mejor_horario_dict = [
            {
                'cursos_grupos': {
                    'curso': sustentacion['cursos_grupos'].curso.nombre,
                    'grupo': sustentacion['cursos_grupos'].grupo.nombre,
                    'profesor': sustentacion['cursos_grupos'].profesor.apellidos_nombres,
                    'semestre': sustentacion['cursos_grupos'].semestre.nombre,
                },
                'estudiante': sustentacion['estudiante'].apellidos_nombres,
                'jurado1': sustentacion['jurado1'].apellidos_nombres if sustentacion['jurado1'] else '',
                'jurado2': sustentacion['jurado2'].apellidos_nombres if sustentacion['jurado2'] else '',
                'asesor': sustentacion['asesor'].apellidos_nombres,
                'titulo': sustentacion['titulo'],
                'fecha': (sustentacion.get('fecha') + timedelta(days=1)).isoformat() if sustentacion.get('fecha') else 'Fecha Inválida',
                'hora_inicio': sustentacion.get('hora_inicio').isoformat() if sustentacion.get('hora_inicio') else '',
                'hora_fin': sustentacion.get('hora_fin').isoformat() if sustentacion.get('hora_fin') else '',
            }
            for sustentacion in mejor_horario
        ]

        mejor_horario_dict.sort(key=lambda x: (
            x['cursos_grupos']['curso'],
            x['fecha'] == 'Fecha Inválida',
            x['fecha'] if x['fecha'] != 'Fecha Inválida' else '',
            x['hora_inicio']
        ))

        request.session['mejor_horario'] = mejor_horario_dict
        request.session['tipo_sustentacion'] = tipo_sustentacion

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'mejor_horario': mejor_horario_dict})
        else:
            return redirect('ejecutar_algoritmo')

    semestre_actual = SemestreAcademico.objects.filter(vigencia=True).first()

    sustentaciones = Sustentacion.objects.all()

    if semestre_id:
        sustentaciones = sustentaciones.filter(cursos_grupos__semestre_id=semestre_id)
    else:
        sustentaciones = sustentaciones.filter(cursos_grupos__semestre=semestre_actual)

    if tipo_sustentacion:
        semanas_sustentacion = Semana_Sustentacion.objects.filter(tipo_sustentacion=tipo_sustentacion, semestre_academico=semestre_actual)
        sustentaciones = sustentaciones.filter(cursos_grupos__curso__in=semanas_sustentacion.values('curso'))

    if curso_grupo_id:
        sustentaciones = sustentaciones.filter(cursos_grupos_id=curso_grupo_id)

    sustentaciones_con_horarios = Horario_Sustentaciones.objects.filter(
        sustentacion__in=sustentaciones
    ).select_related(
        'sustentacion'
    ).order_by(
        'fecha',
        'hora_inicio',
        'sustentacion__cursos_grupos__curso__nombre'
    )

    for horario in sustentaciones_con_horarios:
        horario.fecha_formateada = formatear_fecha(horario.fecha)

    tipo_sustentacion_mostrado = ""
    fechas_sustentaciones = list(sustentaciones_con_horarios.values_list('fecha', flat=True).distinct())
    if fechas_sustentaciones:
        semanas_sustentacion = Semana_Sustentacion.objects.filter(semestre_academico=semestre_actual)
        for semana in semanas_sustentacion:
            if any(fecha and semana.fecha_inicio <= fecha <= semana.fecha_fin for fecha in fechas_sustentaciones):
                tipo_sustentacion_mostrado = semana.tipo_sustentacion
                break

    context = {
        'sustentaciones_con_horarios': sustentaciones_con_horarios,
        'semestres': semestres,
        'cursos_grupos': cursos_grupos,
        'semestre_id': semestre_id,
        'tipo_sustentacion': tipo_sustentacion,
        'curso_grupo_id': curso_grupo_id,
        'tipo_sustentacion_mostrado': tipo_sustentacion_mostrado,
    }

    return render(request, 'admin/ejecutar_algoritmo.html', context)

def formatear_fecha(fecha):
    if fecha:
        return format_date(fecha, format='EEEE, dd/MM/yyyy', locale='es').capitalize()
    return "Por asignar"


@staff_member_required
def guardar_horarios(request):
    if request.method == 'POST':
        mejor_horario = request.session.get('mejor_horario', [])
        if mejor_horario:
            try:
                guardar_horario(mejor_horario)
                messages.success(request, "Horarios guardados exitosamente.")
            except Exception as e:
                messages.error(request, str(e))
                return redirect('ejecutar_algoritmo')
        else:
            messages.error(request, "No hay horarios para guardar.")
        return redirect('ejecutar_algoritmo')
    return redirect('home')



@staff_member_required
def editar_sustentacion(request, sustentacion_id):
    # Obtener la instancia de la sustentación a editar
    sustentacion = get_object_or_404(Sustentacion, pk=sustentacion_id)

    if request.method == 'POST':
        # Crear una instancia del formulario y llenarlo con los datos recibidos
        form = Horario_SustentacionForm(request.POST, instance=sustentacion)
        if form.is_valid():
            # Guardar los datos en la base de datos
            form.save()
            print("Formulario registrado")
            # Mostrar mensaje de éxito
            messages.success(request, "Sustentación guardada exitosamente.")
            # Redirigir al usuario a la página principal o a donde prefieras
            return redirect('ejecutar_algoritmo')
        else:
            print("Errores en el formulario: ", form.errors)
            # Si el formulario no es válido, mostrar los errores
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        # Si la solicitud es GET, crear una instancia del formulario con los datos de la sustentación actual
        form = Horario_SustentacionForm(instance=sustentacion)
        horario = Horario_Sustentaciones.objects.filter(sustentacion=sustentacion).first()
        if horario:
            print("Fecha de la sustentación: ", horario.fecha)
        else:
            print("No se encontró horario para esta sustentación")

    # Renderizar el formulario de edición
    return render(request, 'admin/editar_sustentacion.html', {'form': form})
def index(request):
    return render(request, 'pages/index.html', { 'segment': 'index' })

def billing(request):
    return render(request, 'pages/billing.html', { 'segment': 'billing' })

def tables(request):
    return render(request, 'pages/tables.html', { 'segment': 'tables' })

def vr(request):
    return render(request, 'pages/virtual-reality.html', { 'segment': 'vr' })

def rtl(request):
    return render(request, 'pages/rtl.html', { 'segment': 'rtl' })

def profile(request):
    return render(request, 'pages/profile.html', { 'segment': 'profile' })

# Authentication
class UserLoginView(LoginView):
    template_name = 'accounts/login.html'
    form_class = LoginForm

def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            print('Account created successfully!')
            return redirect('/accounts/login/')
        else:
            print("Register failed!")
    else:
        form = RegistrationForm()

    context = { 'form': form }
    return render(request, 'accounts/register.html', context)

def logout_view(request):
    logout(request)
    return redirect('/accounts/login/')

class UserPasswordResetView(PasswordResetView):
    template_name = 'accounts/password_reset.html'
    form_class = UserPasswordResetForm

class UserPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'accounts/password_reset_confirm.html'
    form_class = UserSetPasswordForm

class UserPasswordChangeView(PasswordChangeView):
    template_name = 'accounts/password_change.html'
    form_class = UserPasswordChangeForm

# Jurados views
@staff_member_required
def jurados_list(request):
    semestres = SemestreAcademico.objects.all().order_by('-vigencia', 'nombre')
    semestre_vigente = SemestreAcademico.objects.filter(vigencia=True).first()

    # Obtener el semestre seleccionado de los parámetros de la solicitud
    selected_semestre_id = request.GET.get('semestre', semestre_vigente.id if semestre_vigente else None)

    # Filtrar los jurados por el semestre seleccionado
    jurados = Semestre_Academico_Profesores.objects.filter(semestre_id=selected_semestre_id)

    query = request.GET.get('q')
    if query:
        jurados = jurados.filter(profesor__apellidos_nombres__icontains=query)

    context = {
        'jurados': jurados,
        'semestres': semestres,
        'selected_semestre_id': int(selected_semestre_id) if selected_semestre_id else None,
        'semestre_vigente': semestre_vigente,
    }

    return render(request, 'admin/jurados_list.html', context)



@staff_member_required
@transaction.atomic
def jurados_create(request):
    if request.method == 'POST':
        form = ProfesorForm(request.POST)
        if form.is_valid():
            profesor = form.save()

            # Obtener los datos adicionales del formulario
            semestre_academico = form.cleaned_data['semestre_academico']
            horas_asesoria_semanal = form.cleaned_data['horas_asesoria_semanal']

            # Crear el registro en Semestre_Academico_Profesores
            Semestre_Academico_Profesores.objects.create(
                semestre=semestre_academico,
                profesor=profesor,
                horas_asesoria_semanal=horas_asesoria_semanal
            )

            messages.success(request, "Jurado creado exitosamente")
            return redirect('jurados_list')
    else:
        form = ProfesorForm()
    return render(request, 'admin/jurados_form.html', {'form': form})

@staff_member_required
def jurados_update(request, pk):
    semestre_academico_profesor = get_object_or_404(Semestre_Academico_Profesores, pk=pk)
    profesor = semestre_academico_profesor.profesor

    if request.method == 'POST':
        form = ProfesorForm(request.POST, instance=profesor)
        if form.is_valid():
            profesor = form.save()

            # Actualizar los datos adicionales del modelo Semestre_Academico_Profesores
            semestre_academico_profesor.semestre = form.cleaned_data['semestre_academico']
            semestre_academico_profesor.horas_asesoria_semanal = form.cleaned_data['horas_asesoria_semanal']
            semestre_academico_profesor.save()

            messages.success(request, "Jurado actualizado exitosamente")
            return redirect('jurados_list')
    else:
        initial_data = {
            'semestre_academico': semestre_academico_profesor.semestre,
            'horas_asesoria_semanal': semestre_academico_profesor.horas_asesoria_semanal,
        }
        form = ProfesorForm(instance=profesor, initial=initial_data)

    return render(request, 'admin/jurados_form.html', {'form': form})
@staff_member_required
def jurados_delete(request, pk):
    jurado = get_object_or_404(Semestre_Academico_Profesores, pk=pk)
    if request.method == 'POST':
        jurado.delete()
        messages.success(request, "Jurado eliminado exitosamente")
        return redirect('jurados_list')
    return render(request, 'admin/jurados_confirm_delete.html', {'jurado': jurado})




from django.db import transaction
from django.contrib.auth.models import User
from App.models import *

@staff_member_required
def jurados_import(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                df = pd.read_excel(file)
                with transaction.atomic():
                    for index, row in df.iterrows():
                        email = row['Email']
                        username = email.split('@')[0]
                        telefono = str(row['Teléfono'])  # Convertir a cadena para la contraseña

                        # Crear usuario en auth_user
                        user, user_created = User.objects.get_or_create(
                            username=username,
                            defaults={
                                'email': email,
                                'first_name': row['Apellidos y nombres'].split(' ')[0],  # Primer nombre
                                'last_name': ' '.join(row['Apellidos y nombres'].split(' ')[1:]),  # Apellidos
                                'is_staff': True,
                                'is_active': True,
                            }
                        )
                        if user_created:
                            user.set_password(telefono)
                            user.save()
                        else:
                            user.email = email
                            user.set_password(telefono)
                            user.first_name = row['Apellidos y nombres'].split(' ')[0]
                            user.last_name = ' '.join(row['Apellidos y nombres'].split(' ')[1:])
                            user.save()

                        # Crear o actualizar el perfil
                        profile, profile_created = Profile.objects.get_or_create(
                            user=user,
                            defaults={'rol': 'P'}
                        )
                        if not profile_created:
                            profile.rol = 'P'
                            profile.save()

                        # Crear profesor
                        profesor, created = Profesor.objects.get_or_create(
                            email=email,
                            defaults={
                                'apellidos_nombres': row['Apellidos y nombres'],
                                'dedicacion': row['Dedicación'],
                                'telefono': telefono,
                                'user': user
                            }
                        )
                        try:
                            semestre = SemestreAcademico.objects.get(nombre=row['Semestre'])
                        except SemestreAcademico.DoesNotExist:
                            messages.error(request, f"El semestre '{row['Semestre']}' no existe en la base de datos.")
                            continue

                        if Semestre_Academico_Profesores.objects.filter(semestre=semestre, profesor=profesor).exists():
                            messages.error(request, f"El profesor '{profesor}' ya está asignado al semestre '{semestre}'.")
                            continue

                        Semestre_Academico_Profesores.objects.update_or_create(
                            semestre=semestre,
                            profesor=profesor,
                            defaults={'horas_asesoria_semanal': row['Horas de asesoría semanal']}
                        )
                messages.success(request, "Profesores importados exitosamente")
            except Exception as e:
                messages.error(request, f"Error al importar el archivo: {e}")
            return redirect('jurados_list')
    else:
        form = ExcelUploadForm()
    return render(request, 'admin/jurados_import.html', {'form': form})


#Sustentacion
# Crear y listar grupos horarios
@staff_member_required
def sustentacion_list(request, semestre_nombre, curso_grupo_nombre, curso_grupo_id):
    curso_grupo = get_object_or_404(Cursos_Grupos, id=curso_grupo_id)
    sustentaciones = Sustentacion.objects.filter(cursos_grupos=curso_grupo)
    query = request.GET.get('q')
    if query:
        sustentaciones = sustentaciones.filter(titulo__icontains=query)
    return render(request, 'admin/sustentacion_list.html', {'sustentaciones': sustentaciones, 'curso_grupo': curso_grupo})

@staff_member_required
def sustentacion_create(request, curso_grupo_id):
    curso_grupo = get_object_or_404(Cursos_Grupos, id=curso_grupo_id)
    if request.method == 'POST':
        form = SustentacionForm(request.POST)
        if form.is_valid():
            sustentacion = form.save(commit=False)
            sustentacion.cursos_grupos = curso_grupo
            sustentacion.save()
            messages.success(request, "Sustentacion creada exitosamente")
            return redirect('sustentacion_list', semestre_nombre=curso_grupo.semestre.nombre, curso_grupo_nombre=curso_grupo.curso.nombre +"("+ curso_grupo.grupo.nombre +")", curso_grupo_id=curso_grupo.id)
    else:
        form = SustentacionForm(initial={'cursos_grupos': curso_grupo})
    return render(request, 'admin/sustentacion_form.html', {'form': form, 'curso_grupo': curso_grupo})

@staff_member_required
def sustentacion_update(request, pk):
    sustentacion = get_object_or_404(Sustentacion, pk=pk)
    curso_grupo = sustentacion.cursos_grupos
    if request.method == 'POST':
        form = SustentacionForm(request.POST, instance=sustentacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Sustentacion actualizada exitosamente")
            return redirect('sustentacion_list', semestre_nombre=curso_grupo.semestre.nombre, curso_grupo_nombre=curso_grupo.curso.nombre +"("+ curso_grupo.grupo.nombre +")", curso_grupo_id=curso_grupo.id)
    else:
        form = SustentacionForm(instance=sustentacion)
    return render(request, 'admin/sustentacion_form.html', {'form': form, 'curso_grupo': curso_grupo})

@staff_member_required
def sustentacion_delete(request, pk):
    sustentacion = get_object_or_404(Sustentacion, pk=pk)
    curso_grupo = sustentacion.cursos_grupos
    if request.method == 'POST':
        sustentacion.delete()
        messages.success(request, "Sustentacion eliminada exitosamente")
        return redirect('sustentacion_list', semestre_nombre=curso_grupo.semestre.nombre, curso_grupo_nombre=curso_grupo.curso.nombre +"("+ curso_grupo.grupo.nombre +")", curso_grupo_id=curso_grupo.id)
    return render(request, 'admin/sustentacion_confirm_delete.html', {'sustentacion': sustentacion})

@staff_member_required
def estudiantes_import(request, curso_grupo_id):
    curso_grupo = get_object_or_404(Cursos_Grupos, id=curso_grupo_id)
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                df = pd.read_excel(file)
                for index, row in df.iterrows():
                    # Crear o actualizar el estudiante
                    estudiante, created = Estudiante.objects.get_or_create(
                        codigo_universitario=row['Código Universitario'],
                        defaults={
                            'apellidos_nombres': row['Apellidos y Nombres'],
                            'email': row['Email'],
                            'telefono': row['Telefono'],
                        }
                    )
                    
                    # Verificar si el estudiante ya está registrado en el mismo curso y grupo
                    if Sustentacion.objects.filter(cursos_grupos=curso_grupo, estudiante=estudiante).exists():
                        messages.warning(request, f"El estudiante '{estudiante.apellidos_nombres}' ya está registrado en este curso y grupo.")
                        continue
                    
                    # Obtener los profesores (jurados y asesor)
                    jurado1 = None
                    jurado2 = None
                    asesor = None

                    if not pd.isnull(row['Jurado 1']):
                        try:
                            jurado1 = Profesor.objects.filter(apellidos_nombres=row['Jurado 1']).first()
                        except Profesor.DoesNotExist:
                            messages.error(request, f"El jurado 1 '{row['Jurado 1']}' no existe.")
                            continue

                    if not pd.isnull(row['Jurado 2']):
                        try:
                            jurado2 = Profesor.objects.filter(apellidos_nombres=row['Jurado 2']).first()
                        except Profesor.DoesNotExist:
                            messages.error(request, f"El jurado 2 '{row['Jurado 2']}' no existe.")
                            continue

                    try:
                        asesor = Profesor.objects.filter(apellidos_nombres=row['Asesor (Jurado 3)']).first()
                    except Profesor.DoesNotExist:
                        messages.error(request, f"El asesor '{row['Asesor (Jurado 3)']}' no existe.")
                        continue


                    try:
                        asesor = Profesor.objects.get(apellidos_nombres=row['Asesor (Jurado 3)'])
                    except Profesor.DoesNotExist:
                        messages.error(request, f"El asesor '{row['Asesor (Jurado 3)']}' no existe.")
                        continue
                    
                    # Crear la sustentacion
                    Sustentacion.objects.create(
                        cursos_grupos=curso_grupo,
                        estudiante=estudiante,
                        jurado1=jurado1,
                        jurado2=jurado2,
                        asesor=asesor,
                        titulo=row['Título de Tesis']
                    )
                messages.success(request, "Estudiantes y sustentaciones importados exitosamente")
            except Exception as e:
                messages.error(request, f"Error al importar el archivo: {e}")
            return redirect('sustentacion_list', semestre_nombre=curso_grupo.semestre.nombre, curso_grupo_nombre=curso_grupo.curso.nombre + "(" + curso_grupo.grupo.nombre + ")", curso_grupo_id=curso_grupo.id)
    else:
        form = ExcelUploadForm()
    return render(request, 'admin/estudiantes_import.html', {'form': form, 'curso_grupo': curso_grupo})





@staff_member_required
def semestre_list(request):
    semestres = SemestreAcademico.objects.all().order_by('-vigencia', 'nombre')
    query = request.GET.get('q')
    if query:
        semestres = semestres.filter(nombre__icontains=query)
    return render(request, 'admin/semestre_list.html', {'semestres': semestres})


@staff_member_required
def semestre_create(request):
    if request.method == 'POST':
        form = SemestreAcademicoForm(request.POST)
        if form.is_valid():
            if form.cleaned_data['vigencia']:
                # Si este semestre se establece como vigente, desactivar los demás
                SemestreAcademico.objects.update(vigencia=False)
            form.save()
            messages.success(request, "Semestre académico creado exitosamente")
            return redirect('semestre_list')
    else:
        form = SemestreAcademicoForm()
    return render(request, 'admin/semestre_form.html', {'form': form})



@staff_member_required
def disponibilidad_list(request):
    if request.user.is_superuser:
        profesores = Profesor.objects.all()
        profesor_id = request.GET.get('profesor_id', profesores.first().id if profesores.exists() else None)
    else:
        profesor = Profesor.objects.filter(user=request.user).first()
        profesor_id = profesor.id if profesor else None
        profesores = [profesor] if profesor else []

    # Realizar la consulta SQL específica
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                assu.semana_inicio, 
                assu.semana_fin
            FROM 
                app_semestreacademico asap 
            JOIN 
                app_semana_sustentacion assu ON assu.semestre_academico_id = asap.id and asap.vigencia = 1
            GROUP BY 
                assu.semana_inicio, 
                assu.semana_fin
            ORDER BY 
                assu.semana_inicio
        """)
        disponibilidades = cursor.fetchall()

    return render(request, 'profesor/disponibilidad_list.html', {
        'disponibilidades': disponibilidades,
        'profesores': profesores,
        'profesor_id': profesor_id,
    })

@staff_member_required
@csrf_exempt
@transaction.atomic
def disponibilidad_create(request):
    if request.user.is_superuser:
        profesor_id = request.GET.get('profesor_id')
        profesor = get_object_or_404(Profesor, id=profesor_id)
    else:
        usuario_logueado = request.user
        profesor = get_object_or_404(Profesor, user=usuario_logueado)
    
    semestre_academico = get_object_or_404(SemestreAcademico, vigencia=True)

    if request.method == 'POST':
        # Obtener semanas de la URL
        semana_inicio = int(request.GET.get('semana_inicio'))
        semana_fin = int(request.GET.get('semana_fin'))

        # Calcular las fechas de inicio y fin de las semanas seleccionadas
        semanas = semestre_academico.calcular_semanas()
        fecha_inicio_semana = semanas[semana_inicio - 1][0]
        fecha_fin_semana = semanas[semana_fin - 1][1]

        # Borrar todas las disponibilidades existentes del profesor para esas semanas
        Profesores_Semestre_Academico.objects.filter(
            profesor=profesor,
            semestre=semestre_academico,
            fecha__range=[fecha_inicio_semana, fecha_fin_semana]
        ).delete()

        # Procesar la solicitud POST para guardar nuevas disponibilidades
        try:
            event_data = json.loads(request.body)
            for event in event_data:
                start = parse_datetime(event['start'])
                end = parse_datetime(event['end'])
                
                # Eliminar cualquier disponibilidad existente para el mismo profesor en la misma fecha y hora
                Profesores_Semestre_Academico.objects.filter(
                    profesor=profesor,
                    semestre=semestre_academico,
                    fecha=start.date(),
                    hora_inicio=start.time(),
                    hora_fin=end.time()
                ).delete()

                # Crear la nueva disponibilidad
                Profesores_Semestre_Academico.objects.create(
                    profesor=profesor,
                    semestre=semestre_academico,
                    fecha=start.date(),
                    hora_inicio=start.time(),
                    hora_fin=end.time()
                )

            return JsonResponse({'status': 'success'})
        except Exception as e:
            print(f"Error al guardar las disponibilidades: {e}")
            return JsonResponse({'status': 'error'})

    # Obtener las disponibilidades existentes
    disponibilidades = Profesores_Semestre_Academico.objects.filter(
        profesor=profesor,
        semestre=semestre_academico
    ).values_list('fecha', 'hora_inicio', 'hora_fin', 'profesor__apellidos_nombres')

    return render(request, 'profesor/disponibilidad_form.html', {
        'disponibilidades': disponibilidades,
        'profesor': profesor
    })

@staff_member_required
def obtener_fechas_min_max(request):
    semana_inicio = request.GET.get('semana_inicio')
    semana_fin = request.GET.get('semana_fin')

    # Asegúrate de convertir semana_inicio y semana_fin en enteros
    semana_inicio = int(semana_inicio)
    semana_fin = int(semana_fin)

    # Obtén las fechas de inicio y fin del semestre
    semestre = SemestreAcademico.objects.get(vigencia=True)  # Ajusta según tus requisitos
    semanas = semestre.calcular_semanas()

    fecha_inicio_min = semanas[semana_inicio - 1][0]
    fecha_fin_max = semanas[semana_fin - 1][1]

    return JsonResponse({
        'fecha_inicio_min': fecha_inicio_min,
        'fecha_fin_max': fecha_fin_max,
    })
    
@staff_member_required
def ver_disponibilidad(request, semana_inicio, semana_fin):
    if request.user.is_superuser:
        profesor_id = request.GET.get('profesor_id')
        profesor = get_object_or_404(Profesor, id=profesor_id)
    else:
        usuario_logueado = request.user
        profesor = get_object_or_404(Profesor, user=usuario_logueado)

    semestre_academico = get_object_or_404(SemestreAcademico, vigencia=True)

    # Calcular las semanas del semestre
    semanas = semestre_academico.calcular_semanas()
    fecha_inicio_semana = semanas[semana_inicio - 1][0]
    fecha_fin_semana = semanas[semana_fin - 1][1]

    # Realizar la consulta SQL específica
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    aps.id, aps.fecha, aps.hora_inicio, aps.hora_fin, 
                    ap.apellidos_nombres
                FROM 
                    app_profesores_semestre_academico aps
                INNER JOIN 
                    app_profesor ap ON aps.profesor_id = ap.id
                INNER JOIN 
                    app_semestreacademico sem ON aps.semestre_id = sem.id
                WHERE 
                    aps.fecha >= %s AND aps.fecha <= %s AND ap.id = %s AND sem.vigencia = true
            """, [fecha_inicio_semana, fecha_fin_semana, profesor.id])
            disponibilidades = cursor.fetchall()
    except DatabaseError as e:
        disponibilidades = []
        print(f"Error en la base de datos: {e}")

    return render(request, 'profesor/ver_disponibilidad.html', {
        'disponibilidades': disponibilidades,
        'semana_inicio': semana_inicio,
        'semana_fin': semana_fin,
        'profesor': profesor
    })

   
    
@staff_member_required
def semestre_update(request, pk):
    semestre = get_object_or_404(SemestreAcademico, pk=pk)
    if request.method == 'POST':
        form = SemestreAcademicoForm(request.POST, instance=semestre)
        if form.is_valid():
            if form.cleaned_data['vigencia']:
                # Si este semestre se establece como vigente, desactivar los demás
                SemestreAcademico.objects.exclude(pk=semestre.pk).update(vigencia=False)
            form.save()
            messages.success(request, "Semestre académico actualizado exitosamente")
            return redirect('semestre_list')
    else:
        form = SemestreAcademicoForm(instance=semestre)
        print(f"Fecha Inicio: {semestre.fecha_inicio}, Fecha Fin: {semestre.fecha_fin}")
    return render(request, 'admin/semestre_form.html', {'form': form, 'semestre': semestre})



@staff_member_required
def semestre_delete(request, pk):
    semestre = get_object_or_404(SemestreAcademico, pk=pk)
    if request.method == 'POST':
        semestre.delete()
        messages.success(request, "Semestre académico eliminado exitosamente")
        return redirect('semestre_list')
    return render(request, 'admin/semestre_confirm_delete.html', {'semestre': semestre})



#grupohorario
@staff_member_required
def grupos_list(request, semestre_nombre):
    semestre = get_object_or_404(SemestreAcademico, nombre=semestre_nombre)
    grupos = Cursos_Grupos.objects.filter(semestre=semestre)
    return render(request, 'admin/grupos_list.html', {'grupos': grupos, 'semestre': semestre})

@staff_member_required
def grupo_create(request, semestre_nombre):
    semestre = get_object_or_404(SemestreAcademico, nombre=semestre_nombre)
    if request.method == 'POST':
        form = CursosGruposForm(request.POST)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.semestre = semestre
            grupo.save()
            messages.success(request, "Grupo horario creado exitosamente")
            return redirect('grupos_list', semestre_nombre=semestre.nombre)
    else:
        form = CursosGruposForm(initial={'semestre': semestre})
    return render(request, 'admin/grupo_form.html', {'form': form, 'semestre': semestre})

@staff_member_required
def grupo_update(request, pk):
    grupo = get_object_or_404(Cursos_Grupos, pk=pk)
    if request.method == 'POST':
        form = CursosGruposForm(request.POST, instance=grupo)
        if form.is_valid():
            form.save()
            messages.success(request, "Grupo horario actualizado exitosamente")
            return redirect('grupos_list', semestre_nombre=grupo.semestre.nombre)
    else:
        form = CursosGruposForm(instance=grupo)
    return render(request, 'admin/grupo_form.html', {'form': form, 'semestre': grupo.semestre})

@staff_member_required
def grupo_delete(request, pk):
    grupo = get_object_or_404(Cursos_Grupos, pk=pk)
    if request.method == 'POST':
        grupo.delete()
        messages.success(request, "Grupo horario eliminado exitosamente")
        return redirect('grupos_list', semestre_nombre=grupo.semestre.nombre)
    return render(request, 'admin/grupo_confirm_delete.html', {'grupo': grupo})

# Semana Sustentacion
def get_semanas(request, semestre_id):
    semestre = get_object_or_404(SemestreAcademico, pk=semestre_id)
    semanas = semestre.calcular_semanas()
    semanas_formateadas = [(str(semana[0]), str(semana[1])) for semana in semanas]
    return JsonResponse(semanas_formateadas, safe=False)
#PARA REPORTES

from django.shortcuts import render
from django.http import HttpResponse
import csv

from django.http import HttpResponse
from django.db import connection
import openpyxl
from openpyxl.utils import get_column_letter

#Reporte 1: Sustentaciones

def reporte_sustentaciones(request):
    fecha = request.GET.get('fecha', '')
    
    # Obtener la lista de semestres
    semestres = SemestreAcademico.objects.all()

    sql = """
    SELECT 
    appc.nombre as curso,
    appg.nombre as grupo,
    appe.codigo_universitario, 
    appe.apellidos_nombres as estudiante,
    appe.email as email_estudiante,
    appe.telefono as telefono_estudiante,
    appc.nombre as curso,
    app1.apellidos_nombres as jurado1,
    app2.apellidos_nombres as jurado2,
    app3.apellidos_nombres as asesor,
    apphs.fecha,
    apphs.hora_inicio,
    apphs.hora_fin,
    apps.titulo
    FROM app_sustentacion apps
    INNER JOIN app_cursos_grupos appcg on apps.cursos_grupos_id=appcg.id
    INNER JOIN app_curso appc on appc.id=appcg.curso_id
    INNER JOIN app_grupo appg on appg.id=appcg.grupo_id
    INNER JOIN app_estudiante appe on appe.id=apps.estudiante_id
    LEFT JOIN app_profesor app1 on app1.id=apps.jurado1_id
    LEFT JOIN app_profesor app2 on app2.id=apps.jurado2_id
    LEFT JOIN app_profesor app3 on app3.id=apps.asesor_id
    INNER JOIN app_horario_sustentaciones apphs on apphs.sustentacion_id=apps.id
    WHERE 1=1
    """

    params = []
    
    if fecha:
        sql += " AND apphs.fecha = %s"
        params.append(fecha)
    
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    data = [
        dict(zip(columns, row))
        for row in rows
    ]

    context = {
        'data': data,
        'fecha': fecha,
    }

    return render(request, 'admin/reporte_sustentaciones.html', context)



def exportar_csv(request):
    fecha = request.GET.get('fecha')
    
    sql = """
    SELECT 
    appc.nombre as curso,
    appg.nombre as grupo,
    appe.codigo_universitario, 
    appe.apellidos_nombres as estudiante,
    appe.email as email_estudiante,
    appe.telefono as telefono_estudiante,
    appc.nombre as curso,
    app1.apellidos_nombres as jurado1,
    app2.apellidos_nombres as jurado2,
    app3.apellidos_nombres as asesor,
    apphs.fecha,
    apphs.hora_inicio,
    apphs.hora_fin,
    apps.titulo
    FROM app_sustentacion apps
    INNER JOIN app_cursos_grupos appcg on apps.cursos_grupos_id=appcg.id
    INNER JOIN app_curso appc on appc.id=appcg.curso_id
    INNER JOIN app_grupo appg on appg.id=appcg.grupo_id
    INNER JOIN app_estudiante appe on appe.id=apps.estudiante_id
    LEFT JOIN app_profesor app1 on app1.id=apps.jurado1_id
    LEFT JOIN app_profesor app2 on app2.id=apps.jurado2_id
    LEFT JOIN app_profesor app3 on app3.id=apps.asesor_id
    INNER JOIN app_horario_sustentaciones apphs on apphs.sustentacion_id=apps.id
    WHERE 1=1
    """

    params = []

    
    if fecha:
        sql += " AND apphs.fecha = %s"
        params.append(fecha)
    
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    # Crear el libro de Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sustentaciones"

    # Escribir los encabezados de columna
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Escribir los datos
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Ajustar el ancho de las columnas
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sustentaciones.xlsx"'
    
    # Guardar el libro en la respuesta
    workbook.save(response)
    
    return response



#Reporte 2: list sustentaciones por docente
from django.http import HttpResponse
from django.db import connection
import openpyxl
from openpyxl.utils import get_column_letter
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

def listar_sustentaciones(request):
    usuario_id = request.user.id
    sql_profesor_id = """
        SELECT id FROM app_profesor WHERE user_id = %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(sql_profesor_id, [usuario_id])
        profesor_id = cursor.fetchone()[0]

    sql = """
        SELECT 
    MAX(apps.id) as id_su,
    appc.nombre AS curso,
    appe.codigo_universitario,
    appe.apellidos_nombres AS estudiante,
    MAX(appss.tipo_sustentacion) as tipo_sustentacion,
    MAX(appss.semana_inicio) as semana_inicio,
    MAX(appss.semana_fin) as semana_fin,
    MAX(apphs.fecha) as fecha,
    MAX(apphs.hora_inicio) as hora_inicio,
    MAX(apphs.hora_fin) as hora_fin,
    MAX(apps.titulo) as titulo
FROM app_semana_sustentacion appss
INNER JOIN app_semestreacademico appsa on appsa.id = appss.semestre_academico_id
INNER JOIN app_curso appc ON appc.id = appss.curso_id
INNER JOIN app_cursos_grupos appcg ON appss.curso_id = appcg.curso_id
INNER JOIN app_sustentacion apps ON apps.cursos_grupos_id = appcg.id
RIGHT JOIN app_horario_sustentaciones apphs ON apphs.sustentacion_id = apps.id
INNER JOIN app_estudiante appe ON appe.id = apps.estudiante_id
LEFT JOIN app_profesor app1 ON app1.id = apps.jurado1_id
LEFT JOIN app_profesor app2 ON app2.id = apps.jurado2_id
LEFT JOIN app_profesor app3 ON app3.id = apps.asesor_id
WHERE appsa.vigencia = TRUE AND (apps.jurado1_id = %s OR apps.jurado2_id = %s OR apps.asesor_id = %s)
GROUP BY appc.nombre, appe.codigo_universitario, appe.apellidos_nombres
      
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, [profesor_id, profesor_id, profesor_id])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    data = [dict(zip(columns, row)) for row in rows]

    return render(request, 'profesor/listar_sustentaciones.html', {'data': data})
 
def exportar_excel_profesor(request):
    usuario_id = request.user.id
    sql_profesor_id = """
        SELECT id FROM app_profesor WHERE user_id = %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(sql_profesor_id, [usuario_id])
        profesor_id = cursor.fetchone()[0]
    
    sql = """
    SELECT 
        appc.nombre AS curso,
        appg.nombre AS grupo,
        appe.codigo_universitario,
        appe.apellidos_nombres AS estudiante,
        apphs.fecha,
        apphs.hora_inicio,
        apps.titulo
    FROM app_sustentacion apps
    INNER JOIN app_estudiante appe ON appe.id = apps.estudiante_id
    LEFT JOIN app_profesor app1 ON app1.id = apps.jurado1_id
    LEFT JOIN app_profesor app2 ON app2.id = apps.jurado2_id
    LEFT JOIN app_profesor app3 ON app3.id = apps.asesor_id
    INNER JOIN app_cursos_grupos appcg ON apps.cursos_grupos_id = appcg.id
    INNER JOIN app_curso appc ON appc.id = appcg.curso_id
    INNER JOIN app_grupo appg ON appg.id = appcg.grupo_id
    INNER JOIN app_horario_sustentaciones apphs ON apphs.sustentacion_id = apps.id
    WHERE apps.jurado1_id = %s OR apps.jurado2_id = %s OR apps.asesor_id = %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(sql, [profesor_id, profesor_id, profesor_id])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    # Crear el libro de Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sustentaciones"

    # Escribir los encabezados de columna
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Escribir los datos
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Ajustar el ancho de las columnas
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sustentaciones_profesor.xlsx"'
    
    # Guardar el libro en la respuesta
    workbook.save(response)
    
    return response

@login_required
def user_profile(request):
    user = request.user
    
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .utils import send_whatsapp_message
from .models import Sustentacion, Cursos_Grupos

@csrf_exempt
def send_bulk_messages_view(request):
    if request.method == 'POST':
        tipo_sustentacion = request.POST.get('tipo_sustentacion')

        # Obtener el semestre vigente
        semestre_vigente = SemestreAcademico.objects.filter(vigencia=True).first()
        if not semestre_vigente:
            return JsonResponse({'status': 'fail', 'message': 'No hay semestre vigente'})

        # Filtrar las sustentaciones según el semestre vigente y el tipo de sustentación
        sustentaciones = Sustentacion.objects.filter(
            cursos_grupos__semestre_id=semestre_vigente.id,
            cursos_grupos__curso__semana_sustentacion__tipo_sustentacion=tipo_sustentacion,
            horario_sustentaciones__fecha__isnull=False
        ).distinct()


        # Ordenar sustentaciones por fecha y hora de inicio
        sustentaciones = sorted(
            sustentaciones, 
            key=lambda s: (s.horario_sustentaciones_set.first().fecha, s.horario_sustentaciones_set.first().hora_inicio)
        )

        def build_message(records):
            message = "Listado de Sustentaciones:\n"
            for record in records:
                curso_grupo = record.cursos_grupos
                horario = record.horario_sustentaciones_set.first()
                message += f"Curso - Grupo: {curso_grupo.curso.nombre} - {curso_grupo.grupo.nombre}\n"
                message += f"Fecha: {horario.fecha}\n"
                message += f"Hora Inicio: {horario.hora_inicio}\n"
                message += f"Hora Fin: {horario.hora_fin}\n"
                message += f"Tipo: {record.cursos_grupos.curso.semana_sustentacion_set.first().tipo_sustentacion}\n\n"
            return message

        # Lista para almacenar los mensajes individuales
        messages_to_send = []

        # Diccionarios para agrupar los mensajes por teléfono
        message_profesores = {}
        message_estudiantes = {}

        # Agrupar las sustentaciones por teléfono
        for sustentacion in sustentaciones:
            horario_sustentacion = sustentacion.horario_sustentaciones_set.first()
            if horario_sustentacion:
                for jurado in [sustentacion.jurado1, sustentacion.jurado2, sustentacion.asesor]:
                    if jurado and jurado.telefono:
                        if jurado.telefono not in message_profesores:
                            message_profesores[jurado.telefono] = []
                        message_profesores[jurado.telefono].append(sustentacion)

                if sustentacion.estudiante.telefono:
                    if sustentacion.estudiante.telefono not in message_estudiantes:
                        message_estudiantes[sustentacion.estudiante.telefono] = []
                    message_estudiantes[sustentacion.estudiante.telefono].append(sustentacion)

        """ # Construir y almacenar mensajes para profesores
        for phone, records in message_profesores.items():
            nombre_profesor = records[0].jurado1.apellidos_nombres if records[0].jurado1.telefono == phone else records[0].jurado2.apellidos_nombres if records[0].jurado2.telefono == phone else records[0].asesor.apellidos_nombres
            message = f"Nombre: {nombre_profesor}\nNo olvidar su horario de sustentación:\n"
            message += build_message(records)
            # Añadir a la lista de mensajes a enviar
            messages_to_send.append((f"+51{phone}", message))

        
        # Construir y almacenar mensajes para estudiantes
        for phone, records in message_estudiantes.items():
            message = f"Nombre: {records[0].estudiante.apellidos_nombres}\nNo olvidar su horario de sustentación:\n"
            message += build_message(records)    
            # Añadir a la lista de mensajes a enviar
            messages_to_send.append((f"+51{phone}", message)) """
            
         # Construir y almacenar mensajes para profesores
        for phone in message_profesores.keys():
            message = "Esta es una plantilla de prueba para Profesores:\n "
            message += """Curso - Grupo: Proyecto de Investigación - X
                          Fecha: 2024-05-18
                          Hora Inicio: 20:00:00
                          Hora Fin: 20:30:00
                          Tipo: PARCIAL""" 
            # Añadir a la lista de mensajes a enviar
            messages_to_send.append((f"+51{phone}", message))

        # Construir y almacenar mensajes para estudiantes
        for phone in message_estudiantes.keys():
            message = "Esta es una plantilla de prueba para Estudiantes:\n "
            message += """Curso - Grupo: Proyecto de Investigación - X
                          Fecha: 2024-05-18
                          Hora Inicio: 20:00:00
                          Hora Fin: 20:30:00
                          Tipo: PARCIAL""" 
            # Añadir a la lista de mensajes a enviar
            messages_to_send.append((f"+51{phone}", message))
            
        # Enviar los mensajes almacenados
        for phone, message in messages_to_send:
            print(f"Enviando mensaje a {phone}:\n{message}")  # Depuración
            result = send_whatsapp_message(phone, message)
            if not result:
                return JsonResponse({'status': 'fail', 'message': f'Error al enviar el mensaje a {phone}'})
            

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'fail', 'message': 'Invalid request method'})

def send_bulk_messages_view2(request):
    if request.method == 'POST':
        number = '+51916702954'
        sustentaciones = Sustentacion.objects.all()
        message = "Listado de Sustentaciones:\n"
        for sustentacion in sustentaciones:
            message = '''Nombre: ALARCON GARCIA ROGER ERNESTO
No olvidar su horario de sustentación:
Curso - Grupo: Proyecto de Investigación - A
Fecha: 2024-05-06
Hora Inicio: 18:30:00
Hora Fin: 19:00:00
Tipo: PARCIAL
Curso - Grupo: Proyecto de Investigación - A
Fecha: 2024-05-07
Hora Inicio: 18:30:00
Hora Fin: 19:00:00
Tipo: PARCIAL
Curso - Grupo: Proyecto de Investigación - A
Fecha: 2024-05-07
Hora Inicio: 19:30:00
Hora Fin: 20:00:00
Tipo: PARCIAL
Curso - Grupo: Proyecto de Investigación - A
Fecha: 2024-05-14
Hora Inicio: 19:30:00
Hora Fin: 20:00:00
Tipo: PARCIAL'''
        result = send_whatsapp_message(number, message)
        
        if result:
            return JsonResponse({'status': 'success', 'result': result})
        else:
            return JsonResponse({'status': 'fail', 'message': 'Error al enviar el mensaje'})
    
    return JsonResponse({'status': 'fail', 'message': 'Invalid request method'})

@login_required
def listar_compensacion_horas(request):
    usuario_id = request.user.id

    sql_profesor_id = """
        SELECT id FROM app_profesor WHERE user_id = %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(sql_profesor_id, [usuario_id])
        profesor_id = cursor.fetchone()[0]

    sql = """
    SELECT 
    MAX(appss.compensan_horas) AS compensan_horas,
    MAX(appss.duracion_sustentacion) AS duracion_sustentacion,
    appc.nombre AS curso,
    appe.codigo_universitario,
    appe.apellidos_nombres AS estudiante,
    MAX(appss.semana_inicio) AS semana_inicio,
    MAX(appss.semana_fin) AS semana_fin,
    MAX(apphs.fecha) AS fecha,
    MAX(apphs.hora_inicio) AS hora_inicio,
    MAX(apphs.hora_fin) AS hora_fin,
    MAX(apps.titulo) AS titulo
FROM app_semana_sustentacion appss
INNER JOIN app_semestreacademico appsa ON appsa.id = appss.semestre_academico_id
INNER JOIN app_curso appc ON appc.id = appss.curso_id
INNER JOIN app_cursos_grupos appcg ON appss.curso_id = appcg.curso_id
INNER JOIN app_sustentacion apps ON apps.cursos_grupos_id = appcg.id
RIGHT JOIN app_horario_sustentaciones apphs ON apphs.sustentacion_id = apps.id
INNER JOIN app_estudiante appe ON appe.id = apps.estudiante_id
LEFT JOIN app_profesor app1 ON app1.id = apps.jurado1_id
LEFT JOIN app_profesor app2 ON app2.id = apps.jurado2_id
LEFT JOIN app_profesor app3 ON app3.id = apps.asesor_id
WHERE appss.compensan_horas = TRUE 
    AND (apps.jurado1_id = %s OR apps.jurado2_id = %s OR apps.asesor_id = %s) 
    AND appsa.vigencia = TRUE
GROUP BY appc.nombre, appe.codigo_universitario, appe.apellidos_nombres
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, [profesor_id, profesor_id, profesor_id])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    data = [dict(zip(columns, row)) for row in rows]

    # Calcular el total de duración de sustentación
    total_duracion_sustentacion = sum(row['duracion_sustentacion'] for row in data)

    return render(request, 'profesor/listar_compensacion_horas.html', {'data': data, 'total_duracion_sustentacion': total_duracion_sustentacion})


from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

def exportar_excel_compensacion_horas(request):
    usuario_id = request.user.id

    sql_profesor_id = """
        SELECT id FROM app_profesor WHERE user_id = %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(sql_profesor_id, [usuario_id])
        profesor_id = cursor.fetchone()[0]

    sql = """
    SELECT 
        appss.compensan_horas,
        appss.duracion_sustentacion,
        appc.nombre AS curso,
        appe.codigo_universitario,
        appe.apellidos_nombres AS estudiante,
        apphs.fecha,
        appss.semana_inicio,
        appss.semana_fin,
        apphs.hora_inicio,
        apphs.hora_fin,
        apps.titulo
    FROM app_semana_sustentacion appss
    INNER JOIN app_semestreacademico appsa on appsa.id=appss.semestre_academico_id
    INNER JOIN app_curso appc ON appc.id = appss.curso_id
    INNER JOIN app_cursos_grupos appcg ON appss.curso_id = appcg.curso_id
    INNER JOIN app_sustentacion apps ON apps.cursos_grupos_id = appcg.id
    INNER JOIN app_horario_sustentaciones apphs ON apphs.sustentacion_id = apps.id
    INNER JOIN app_estudiante appe ON appe.id = apps.estudiante_id
    LEFT JOIN app_profesor app1 ON app1.id = apps.jurado1_id
    LEFT JOIN app_profesor app2 ON app2.id = apps.jurado2_id
    LEFT JOIN app_profesor app3 ON app3.id = apps.asesor_id
    WHERE appss.compensan_horas = TRUE AND (apps.jurado1_id = %s OR apps.jurado2_id = %s OR apps.asesor_id = %s) AND appsa.vigencia=TRUE
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, [profesor_id, profesor_id, profesor_id])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    # Crear el libro de Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Compensacion Horas"

    # Escribir los encabezados de columna
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Escribir los datos
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Ajustar el ancho de las columnas
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="compensacion_horas_profesor.xlsx"'
    
    # Guardar el libro en la respuesta
    workbook.save(response)
    
    return response

from django.shortcuts import render
from django.http import HttpResponse
from django.db import connection
import csv
from .models import Profesor

def lista_sustentaciones(request):
    query = """
        SELECT 
    MAX(appss.compensan_horas) as compensan_horas,
    MAX(appss.duracion_sustentacion) as duracion_sustentacion,
    appc.nombre AS curso,
    MAX(appss.semana_inicio) as semana_inicio,
    MAX(appss.semana_fin) as semana_fin,
    MAX(apphs.fecha) as fecha,
    MAX(apphs.hora_inicio) as hora_inicio,
    MAX(apphs.hora_fin) as hora_fin,
    MAX(apps.titulo) as titulo,
    MAX(app1.apellidos_nombres) as jurado_1,
    MAX(app2.apellidos_nombres) as jurado_2,
    MAX(app3.apellidos_nombres) as asesor
FROM app_semana_sustentacion appss
INNER JOIN app_semestreacademico appsa ON appsa.id = appss.semestre_academico_id
INNER JOIN app_curso appc ON appc.id = appss.curso_id
INNER JOIN app_cursos_grupos appcg ON appss.curso_id = appcg.curso_id
INNER JOIN app_sustentacion apps ON apps.cursos_grupos_id = appcg.id
LEFT JOIN app_horario_sustentaciones apphs ON apphs.sustentacion_id = apps.id
INNER JOIN app_estudiante appe ON appe.id = apps.estudiante_id
LEFT JOIN app_profesor app1 ON app1.id = apps.jurado1_id
LEFT JOIN app_profesor app2 ON app2.id = apps.jurado2_id
LEFT JOIN app_profesor app3 ON app3.id = apps.asesor_id
WHERE appsa.vigencia = TRUE
GROUP BY appc.nombre, appe.codigo_universitario, appe.apellidos_nombres
    """
    
    params = []
    
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        sustentaciones = [dict(zip(columns, row)) for row in cursor.fetchall()]

    if 'export' in request.GET:
        return exportar_excel(sustentaciones)
    
    return render(request, 'admin/lista_sustentaciones.html', {'sustentaciones': sustentaciones, 'profesores': Profesor.objects.all()})

from openpyxl import Workbook
from django.http import HttpResponse

def exportar_excel(sustentaciones):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sustentaciones"
    
    # Escribir encabezados
    headers = [
        'Compensan Horas', 'Duración Sustentación', 'Curso', 'Semana Inicio', 'Semana Fin', 'Fecha',
        'Hora Inicio', 'Hora Fin', 'Título', 'Jurado 1', 'Jurado 2', 'Asesor'
    ]
    ws.append(headers)
    
    # Escribir datos
    for s in sustentaciones:
        row = [
            'Sí compensa' if s['compensan_horas'] else 'No compensa',
            s['duracion_sustentacion'], s['curso'], s['semana_inicio'], s['semana_fin'],
            s['fecha'], s['hora_inicio'], s['hora_fin'], s['titulo'],
            s['jurado_1'], s['jurado_2'], s['asesor']
        ]
        ws.append(row)
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sustentaciones.xlsx"'
    wb.save(response)
    
    return response

